import sys
import openpyxl

def compare_excels(file1: str, file2: str):
    wb1 = openpyxl.load_workbook(file1, data_only=True)
    wb2 = openpyxl.load_workbook(file2, data_only=True)

    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)

    # Проверяем, есть ли несовпадающие листы
    only_in_file1 = sheets1 - sheets2
    only_in_file2 = sheets2 - sheets1
    if only_in_file1:
        print(f"Листы только в {file1}: {only_in_file1}")
    if only_in_file2:
        print(f"Листы только в {file2}: {only_in_file2}")

    # Сравниваем листы, которые есть в обоих файлах
    common_sheets = sheets1 & sheets2
    for sheet_name in common_sheets:
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]

        max_row = max(ws1.max_row, ws2.max_row)
        max_col = max(ws1.max_column, ws2.max_column)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val1 = ws1.cell(row=row, column=col).value
                val2 = ws2.cell(row=row, column=col).value
                if val1 != val2:
                    cell_name = openpyxl.utils.get_column_letter(col) + str(row)
                    print(
                        f"[{sheet_name}] Различие в ячейке {cell_name}: "
                        f"{file1}='{val1}' | {file2}='{val2}'"
                    )

def main(path1: str, path2: str):
    compare_excels(path1, path2)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Использование: python compare_excels.py file1.xlsx file2.xlsx")
    else:
        main(sys.argv[1], sys.argv[2])
